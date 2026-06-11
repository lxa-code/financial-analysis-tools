import xlrd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 关键指标的行号映射（基于利润预算表.xls的结构）
KEY_INDICATORS = {
    '营业总收入': 3,  # 行4（0-based index=3）
    '营业收入': 4,     # 行5
    '营业总成本': 8,     # 行9
    '营业成本': 9,       # 行10
    '营业利润': 39,      # 行40
    '利润总额': 52,      # 行53
    '净利润': 57,        # 行58
    '归属于母公司净利润': 58,  # 行59
}

def extract_unit_name(ws):
    """从工作表提取单位名称"""
    cell_value = ws.cell_value(1, 0)  # A2单元格
    # 格式："编制单位：河南城市发展投资有限公司（合并）_人民币"
    match = re.search(r'编制单位：(.+?)_人民币', cell_value)
    if match:
        return match.group(1)
    # 如果格式不匹配，尝试直接提取
    return cell_value.replace('编制单位：', '').replace('_人民币', '').strip()

def extract_indicator_value(ws, row_idx, col_idx):
    """提取指标数值，处理空值"""
    try:
        value = ws.cell_value(row_idx, col_idx)
        if value == '' or value is None:
            return 0.0
        return float(value)
    except:
        return 0.0

def process_budget_file(file_path):
    """处理单个预算文件，提取关键指标"""
    try:
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        
        # 提取单位名称
        unit_name = extract_unit_name(ws)
        
        # 提取关键指标
        indicators = {}
        for name, row_idx in KEY_INDICATORS.items():
            if row_idx < ws.nrows:
                indicators[name] = {
                    '2025实际': extract_indicator_value(ws, row_idx, 1),
                    '2026预算': extract_indicator_value(ws, row_idx, 2),
                    '增减率': extract_indicator_value(ws, row_idx, 3)
                }
        
        return unit_name, indicators
    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {str(e)}")
        return None, None

def create_summary_workbook(data_dict):
    """创建汇总Excel工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.title = "预算数据汇总"
    
    # 设置列标题
    headers = ['公司名称']
    for name in KEY_INDICATORS.keys():
        headers.extend([f'{name}(2025实际)', f'{name}(2026预算)', f'{name}(增减率)'])
    
    # 写入标题行
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据行
    row_idx = 2
    for unit_name, indicators in data_dict.items():
        col_idx = 1
        ws.cell(row=row_idx, column=col_idx, value=unit_name)
        
        for ind_name in KEY_INDICATORS.keys():
            if ind_name in indicators:
                ws.cell(row=row_idx, column=col_idx+1, value=indicators[ind_name]['2025实际'])
                ws.cell(row=row_idx, column=col_idx+2, value=indicators[ind_name]['2026预算'])
                ws.cell(row=row_idx, column=col_idx+3, value=indicators[ind_name]['增减率'])
            col_idx += 3
        
        row_idx += 1
    
    # 调整列宽
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            try:
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def main():
    """主函数"""
    # 获取所有预算文件
    budget_dir = r'c:\Users\lecoo\CodeBuddy\李新安工作空间\城发投资工作专区\02_财务管理\预算数据'
    budget_files = [f for f in os.listdir(budget_dir) if f.startswith('利润预算表') and f.endswith('.xls')]
    budget_files.sort()
    
    print(f'找到 {len(budget_files)} 个预算文件')
    
    # 处理所有文件
    all_data = {}
    for file_name in budget_files:
        file_path = os.path.join(budget_dir, file_name)
        print(f'处理: {file_name}')
        
        unit_name, indicators = process_budget_file(file_path)
        if unit_name and indicators:
            all_data[unit_name] = indicators
            print('  OK: %s' % unit_name)
    
    # 创建汇总工作簿
    print(f'\n共处理 {len(all_data)} 个公司的预算数据')
    
    wb = create_summary_workbook(all_data)
    
    # 保存文件
    output_path = os.path.join(budget_dir, '预算数据汇总表.xlsx')
    wb.save(output_path)
    print(f'\n汇总文件已保存: {output_path}')

if __name__ == '__main__':
    main()
