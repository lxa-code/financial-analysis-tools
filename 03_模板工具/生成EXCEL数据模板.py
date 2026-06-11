#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成EXCEL数据填充模板
包含WORD文字模板中所有需要填充的数据字段
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def setup_header(ws, title, headers, row=1):
    """设置表头和样式"""
    # 标题行
    ws.merge_cells(f'A{row}:{get_column_letter(len(headers))}{row}')
    cell = ws[f'A{row}']
    cell.value = title
    cell.font = Font(name='黑体', size=14, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # 表头行
    header_row = row + 1
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=i, value=header)
        cell.font = Font(name='黑体', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    return header_row

def add_data_row(ws, row, data, header_row):
    """添加数据行"""
    for i, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=i, value=value)
        cell.font = Font(name='仿宋_GB2312', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

def generate_excel_template():
    """生成EXCEL数据填充模板"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认sheet
    
    # ========== Sheet1: 参数设置 ==========
    ws = wb.create_sheet('参数设置')
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    
    setup_header(ws, '参数设置', ['参数名称', '参数值', '说明'], row=1)
    params = [
        ['YYYY', '2026', '年份（如2026）'],
        ['MM', '04', '月份（如04表示4月）'],
        ['公司名称', '城发投资', '公司全称'],
        ['报告期间', '2026年4月', '报告期间描述']
    ]
    for i, param in enumerate(params, 1):
        add_data_row(ws, i+2, param, 2)
    
    # ========== Sheet2: 关键指标整体 ==========
    ws = wb.create_sheet('关键指标_整体')
    headers = ['指标名称', '本期值（万元）', '上年同期（万元）', '同比增减（万元）', '同比增减率（%）']
    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [20, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '关键指标完成情况 - 整体', headers, row=1)
    data = [
        ['利润总额', '', '', '', ''],
        ['管理费用', '', '', '', ''],
        ['回款金额', '', '', '', '']
    ]
    for i, row_data in enumerate(data, 1):
        add_data_row(ws, i+header_row, row_data, header_row)
    
    # ========== Sheet3: 关键指标本部 ==========
    ws = wb.create_sheet('关键指标_本部')
    headers = ['指标名称', '本期值（万元）', '上年同期（万元）', '同比增减（万元）', '同比增减率（%）']
    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [20, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '关键指标完成情况 - 本部', headers, row=1)
    data = [
        ['利润总额', '', '', '', ''],
        ['管理费用', '', '', '', ''],
        ['资金余额', '', '', '', '-']
    ]
    for i, row_data in enumerate(data, 1):
        add_data_row(ws, i+header_row, row_data, header_row)
    
    # ========== Sheet4: 盈利情况整体 ==========
    ws = wb.create_sheet('盈利情况_整体')
    headers = ['项目', '本期值（万元）', '上年同期（万元）', '同比增减（万元）', '同比增减率（%）']
    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [20, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '盈利情况 - 整体', headers, row=1)
    data = [
        ['营业收入', '', '', '', ''],
        ['营业成本', '', '', '', ''],
        ['利润总额', '', '', '', ''],
        ['净利润', '', '', '', '']
    ]
    for i, row_data in enumerate(data, 1):
        add_data_row(ws, i+header_row, row_data, header_row)
    
    # ========== Sheet5: 盈利情况本部 ==========
    ws = wb.create_sheet('盈利情况_本部')
    headers = ['项目', '本期值（万元）', '上年同期（万元）', '同比增减（万元）', '同比增减率（%）']
    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [20, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '盈利情况 - 本部', headers, row=1)
    data = [
        ['营业收入', '', '', '', ''],
        ['利润总额', '', '', '', '']
    ]
    for i, row_data in enumerate(data, 1):
        add_data_row(ws, i+header_row, row_data, header_row)
    
    # ========== Sheet6: 下属企业盈利 ==========
    ws = wb.create_sheet('下属企业盈利')
    headers = ['企业名称', '营业收入（万元）', '利润总额（万元）', '净利润（万元）', '同比增减率（%）']
    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [25, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '下属企业盈利情况', headers, row=1)
    # 预留10行
    for i in range(1, 11):
        add_data_row(ws, i+header_row, ['', '', '', '', ''], header_row)
    
    # ========== Sheet7: 资产负债整体 ==========
    ws = wb.create_sheet('资产负债_整体')
    headers = ['项目', '本期值（万元）', '上年末（万元）', '增减额（万元）', '增减率（%）']
    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [20, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '资产负债情况 - 整体', headers, row=1)
    data = [
        ['资产总计', '', '', '', ''],
        ['负债合计', '', '', '', ''],
        ['所有者权益', '', '', '', ''],
        ['资产负债率（%）', '', '', '-', '-']
    ]
    for i, row_data in enumerate(data, 1):
        add_data_row(ws, i+header_row, row_data, header_row)
    
    # ========== Sheet8: 资产负债本部 ==========
    ws = wb.create_sheet('资产负债_本部')
    headers = ['项目', '本期值（万元）', '上年末（万元）', '增减额（万元）', '增减率（%）']
    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [20, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '资产负债情况 - 本部', headers, row=1)
    data = [
        ['资产总计', '', '', '', ''],
        ['负债合计', '', '', '', ''],
        ['资产负债率（%）', '', '', '-', '-']
    ]
    for i, row_data in enumerate(data, 1):
        add_data_row(ws, i+header_row, row_data, header_row)
    
    # ========== Sheet9: 资金情况整体 ==========
    ws = wb.create_sheet('资金情况_整体')
    headers = ['项目', '本期值（万元）', '说明']
    for col, width in zip(['A', 'B', 'C'], [20, 18, 30]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '资金情况 - 整体', headers, row=1)
    data = [
        ['资金余额', '', '整体资金余额'],
        ['货币资金', '', '其中：货币资金'],
        ['其他流动资金', '', '其他流动资金']
    ]
    for i, row_data in enumerate(data, 1):
        add_data_row(ws, i+header_row, row_data, header_row)
    
    # ========== Sheet10: 资金情况本部 ==========
    ws = wb.create_sheet('资金情况_本部')
    headers = ['项目', '本期值（万元）', '上月值（万元）', '环比增减（万元）', '说明']
    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [20, 18, 18, 18, 30]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '资金情况 - 本部', headers, row=1)
    data = [
        ['资金余额', '', '', '', '本部资金余额'],
    ]
    for i, row_data in enumerate(data, 1):
        add_data_row(ws, i+header_row, row_data, header_row)
    
    # ========== Sheet11: 归集情况 ==========
    ws = wb.create_sheet('归集情况')
    headers = ['企业名称', '资金余额（万元）', '归集金额（万元）', '归集率（%）']
    for col, width in zip(['A', 'B', 'C', 'D'], [25, 18, 18, 15]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '资金归集情况', headers, row=1)
    # 预留10行
    for i in range(1, 11):
        add_data_row(ws, i+header_row, ['', '', '', ''], header_row)
    
    # ========== Sheet12: 投资情况 ==========
    ws = wb.create_sheet('投资情况')
    headers = ['项目', '金额（万元）', '说明']
    for col, width in zip(['A', 'B', 'C'], [25, 18, 30]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '投资情况', headers, row=1)
    data = [
        ['股权投资总额', '', '截至本月末股权投资总额'],
        ['本年新增股权投资', '', '本年新增股权投资金额'],
        ['债权投资总额', '', '截至本月末债权投资总额'],
        ['委托贷款总额', '', '其中：委托贷款总额']
    ]
    for i, row_data in enumerate(data, 1):
        add_data_row(ws, i+header_row, row_data, header_row)
    
    # ========== Sheet13: 预算执行情况 ==========
    ws = wb.create_sheet('预算执行')
    headers = ['项目', '年度预算（万元）', '实际完成（万元）', '执行率（%）', '序时进度（%）']
    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [20, 18, 18, 15, 15]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '预算执行情况', headers, row=1)
    data = [
        ['利润总额', '', '', '', ''],
        ['管理费用', '', '', '', ''],
        ['回款金额', '', '', '', '']
    ]
    for i, row_data in enumerate(data, 1):
        add_data_row(ws, i+header_row, row_data, header_row)
    
    # ========== Sheet14: 预算计划 ==========
    ws = wb.create_sheet('预算计划')
    headers = ['项目', '待完成金额（万元）', '说明']
    for col, width in zip(['A', 'B', 'C'], [20, 18, 30]):
        ws.column_dimensions[col].width = width
    
    header_row = setup_header(ws, '预算计划情况', headers, row=1)
    data = [
        ['利润总额待完成', '', '后续月份需完成利润总额'],
        ['管理费用待完成', '', '后续月份管理费用控制目标']
    ]
    for i, row_data in enumerate(data, 1):
        add_data_row(ws, i+header_row, row_data, header_row)
    
    # ========== 保存文件 ==========
    output_path = r'c:\Users\lecoo\CodeBuddy\20260425100224\城发投资月度财务数据填充模板.xlsx'
    wb.save(output_path)
    print(f'EXCEL数据填充模板已生成：{output_path}')
    print()
    print('模板包含14个sheet：')
    print('  1. 参数设置 - YYYY、MM等参数')
    print('  2. 关键指标_整体 - 利润总额、管理费用、回款金额')
    print('  3. 关键指标_本部 - 利润总额、管理费用、资金余额')
    print('  4. 盈利情况_整体 - 营业收入、成本、利润、净利润')
    print('  5. 盈利情况_本部 - 营业收入、利润总额')
    print('  6. 下属企业盈利 - 各企业盈利数据（预留10行）')
    print('  7. 资产负债_整体 - 资产、负债、权益、资产负债率')
    print('  8. 资产负债_本部 - 资产、负债、资产负债率')
    print('  9. 资金情况_整体 - 资金余额、货币资金、其他流动资金')
    print('  10. 资金情况_本部 - 资金余额及环比')
    print('  11. 归集情况 - 各企业资金归集数据（预留10行）')
    print('  12. 投资情况 - 股权、债权投资数据')
    print('  13. 预算执行 - 利润、管理费用、回款预算执行')
    print('  14. 预算计划 - 后续月份待完成金额')
    print()
    print('使用方法：')
    print('  1. 在EXCEL模板中填写各sheet的数据')
    print('  2. 使用Python脚本读取EXCEL数据')
    print('  3. 将数据填充到WORD文字模板的占位符中')
    print('  4. 生成最终的月度财务情况说明文档')

if __name__ == '__main__':
    generate_excel_template()
