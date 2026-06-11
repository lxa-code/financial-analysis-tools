import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'd:\BaiduSyncdisk\CODE\城发投资工作专区\10_日常工作\综合管理部月报\月度重点计划模板 6月 0601 - 综合管理部.xlsx', data_only=True)
ws = wb['5月度重点工作总结']

headers = []
for cell in ws[2]:
    headers.append(str(cell.value) if cell.value is not None else '')

rows_data = []
for r in range(3, ws.max_row+1):
    seq = ws.cell(row=r, column=1).value
    if seq is None:
        continue
    row_dict = {}
    for c in range(1, len(headers)+1):
        val = ws.cell(row=r, column=c).value
        row_dict[headers[c-1]] = str(val) if val is not None else ''
    rows_data.append(row_dict)

issues = []

for i, rd in enumerate(rows_data):
    seq = rd.get('序号', '')
    task = rd.get('一级任务', '')
    sub2 = rd.get('二级任务', '')
    sub3 = rd.get('三级任务', '')
    strategy = rd.get('关联战略目标', '')
    leader = rd.get('交办领导', '')
    responsible = rd.get('责任人/部门', '')
    deadline = rd.get('计划完成时间', '')
    plan_may = rd.get('5月份计划完成成果', '')
    actual_may = rd.get('5月份完成成果', '')
    progress = rd.get('当前进度(%)', '')
    risk = rd.get('风险与问题', '')
    evidence = rd.get('成果支撑/链接', '')
    priority = rd.get('优先级', '')
    source = rd.get('交办来源', '')
    category = rd.get('事项类别', '')
    week_plan = rd.get('周分解计划（关键节点）', '')

    # 1. 检查：序号重复
    # (handled below)

    # 2. 检查：计划完成时间和5月计划完成成果为空
    if not deadline.strip() and seq not in ['4', '5']:
        issues.append(f"[Row#{i+1}, 序号{seq}] 计划完成时间为空，需补充")
    if not plan_may.strip() and seq not in ['4', '5']:
        issues.append(f"[Row#{i+1}, 序号{seq}] 5月份计划完成成果为空，需补充")

    # 3. 检查：5月份完成成果 = 5月份计划完成成果（逐字复制，没写实际完成情况）
    if plan_may.strip() and actual_may.strip() and plan_may.strip() == actual_may.strip():
        issues.append(f"[Row#{i+1}, 序号{seq}] 5月份完成成果与5月份计划完成成果完全一致，疑似直接复制计划未填写实际完成情况")

    # 4. 检查：进度为0或异常低，但5月完成成果有内容
    try:
        pct_str = progress.strip().replace('%', '').replace('，', ',').replace('、', ',')
        # Handle multi-part progress like "1.100%\n2.50%\n3.20%"
        parts = [p.strip() for p in pct_str.split('\n') if p.strip()]
        if parts:
            # Try parsing each part as "N.XX%" or just "XX%"
            pcts = []
            for p in parts:
                # Remove leading number prefix like "1."
                if '.' in p and p.split('.')[0].isdigit():
                    p = p.split('.', 1)[1]
                p = p.strip().replace('%', '')
                try:
                    pcts.append(float(p))
                except:
                    pass
            if pcts and all(p == 0 for p in pcts) and actual_may.strip():
                issues.append(f"[Row#{i+1}, 序号{seq}] 当前进度为0%，但5月份完成成果有内容，进度与成果不匹配")
    except:
        pass

    # 5. 检查：进度100%但5月完成成果为空或为计划描述
    try:
        pct_str2 = progress.strip().replace('%', '').replace('，', ',').replace('、', ',')
        parts2 = [p.strip() for p in pct_str2.split('\n') if p.strip()]
        if parts2:
            pcts2 = []
            for p in parts2:
                if '.' in p and p.split('.')[0].isdigit():
                    p = p.split('.', 1)[1]
                p = p.strip().replace('%', '')
                try:
                    pcts2.append(float(p))
                except:
                    pass
            if pcts2 and all(p == 100 for p in pcts2) and (not actual_may.strip() or actual_may.strip() == plan_may.strip()):
                issues.append(f"[Row#{i+1}, 序号{seq}] 进度显示100%，但5月完成成果为空或与计划一致，疑未更新实际成果")
    except:
        pass

    # 6. 检查：风险描述是"困难"而非"未完成的后果"
    difficulty_keywords = ['时间紧张', '人员不足', '工作量', '理解', '掌握', '推广', '使用不充分', '作用未发挥', '不及时', '不到位', '收到政策影响', '无法跟踪', '修订不及时']
    for kw in difficulty_keywords:
        if kw in risk:
            issues.append(f"[Row#{i+1}, 序号{seq}] 风险描述含'{kw}'，像是工作困难而非未完成的后果，建议改为'如XX不完成，将导致XX'的表述")

    # 7. 检查：责任人/部门为空（但有序号和任务内容）
    if not responsible.strip():
        issues.append(f"[Row#{i+1}, 序号{seq}] 责任人/部门为空，需指定")

    # 8. 检查：P0任务进度偏低
    if priority.strip() == 'P0':
        try:
            pct_str3 = progress.strip().replace('%', '').replace('，', ',').replace('、', ',')
            parts3 = [p.strip() for p in pct_str3.split('\n') if p.strip()]
            if parts3:
                for p in parts3:
                    if '.' in p and p.split('.')[0].isdigit():
                        p = p.split('.', 1)[1]
                    p = p.strip().replace('%', '')
                    try:
                        v = float(p)
                        if v < 30 and v > 0:
                            issues.append(f"[Row#{i+1}, 序号{seq}] P0任务子项进度仅{v}%，需关注风险")
                    except:
                        pass
        except:
            pass

    # 9. 检查：5月份完成成果内容描述不像是5月实际完成
    future_keywords = ['6月', '6月底', '六月底', '6月份', '六月', '谋划', '7月底']
    for kw in future_keywords:
        if kw in actual_may:
            issues.append(f"[Row#{i+1}, 序号{seq}] 5月完成成果中出现'{kw}'，这是未来时间，不应出现在5月总结中")

    # 10. 检查：成果支撑太笼统
    if evidence.strip() in ['方案', '报告', '台账', '工作方案', '问题整改', '发展党员', '完成学习教育']:
        issues.append(f"[Row#{i+1}, 序号{seq}] 成果支撑'{evidence}'过于笼统，建议写具体文件名称")

    # 11. 检查：序号重复
    # Will check separately

    # 12. 检查：周分解计划中有空行占位符
    if '...' in week_plan or '…' in week_plan:
        issues.append(f"[Row#{i+1}, 序号{seq}] 周分解计划中含有'...'占位符，未填写具体内容")

    # 13. 检查：进度用"0.021"这种异常值
    if '0.021' in progress or '0.02' in progress:
        issues.append(f"[Row#{i+1}, 序号{seq}] 进度值异常({progress})，疑似应为21%而非0.021%")

    # 14. 检查：5月完成成果编号跳跃（如从2跳到4）
    if '4.' in actual_may and '3.' not in actual_may:
        issues.append(f"[Row#{i+1}, 序号{seq}] 5月完成成果编号从2跳到4，缺少第3项，需确认")

    # 15. 检查：计划完成时间是数字而非日期
    if deadline.strip().isdigit():
        issues.append(f"[Row#{i+1}, 序号{seq}] 计划完成时间为数字'{deadline}'，疑似Excel日期序列值未正确格式化")

print("=" * 60)
print("5月重点工作总结 — 逻辑检查报告")
print("=" * 60)
print()

if issues:
    print(f"共发现 {len(issues)} 个问题：\n")
    for idx, issue in enumerate(issues, 1):
        print(f"{idx}. {issue}")
else:
    print("未发现明显逻辑问题。")

print()
print("=" * 60)

# Also check for duplicate sequence numbers
seq_counts = {}
for rd in rows_data:
    s = rd.get('序号', '')
    seq_counts[s] = seq_counts.get(s, 0) + 1

dups = {k: v for k, v in seq_counts.items() if v > 1}
if dups:
    print(f"\n序号重复：{dups}")
